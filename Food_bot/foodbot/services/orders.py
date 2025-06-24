from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Tuple

from openpyxl import Workbook
from sqlmodel import select
from sqlalchemy.orm import selectinload
from zoneinfo import ZoneInfo

from foodbot.database import get_session
from foodbot.models import Order, Meal, Canteen, MenuItem, DailyMenu
from foodbot.config import settings


def _local_day_bounds_to_utc(date_local: datetime) -> tuple[datetime, datetime]:
    """Преобразует границы локального дня в UTC-naive время.

    Предполагаем, что *created_at* хранится в UTC-naive.
    Берём смещение системы `datetime.now() - datetime.utcnow()` и
    сдвигаем границы на это значение, получая UTC-эквивалент.
    """
    offset = datetime.now() - datetime.utcnow()
    start_local = date_local.replace(hour=0, minute=0, second=0, microsecond=0)
    start_utc = start_local - offset
    end_utc = start_utc + timedelta(days=1)
    return start_utc, end_utc


async def collect_orders_for_day(date_local: datetime) -> list[Order]:
    """Возвращает заказы за *локальный* день `date_local`.

    created_at хранится в UTC-naive, поэтому сначала переводим границы
    локального дня в UTC, а потом фильтруем.
    """
    start_dt, end_dt = _local_day_bounds_to_utc(date_local)

    async with get_session() as session:
        stmt = (
            select(Order)
            .where(
                Order.created_at >= start_dt,
                Order.created_at < end_dt,
                Order.is_final == True,  # noqa: E712
            )
            .options(
                selectinload(Order.user),
                selectinload(Order.meal).selectinload(Meal.canteen),
            )
        )
        return list(await session.exec(stmt))


async def build_report(date: datetime) -> Tuple[str, BytesIO]:
    """Сформировать Excel-отчёт и вернуть (filename, bytes_io)."""
    orders = await collect_orders_for_day(date)
    if not orders:
        raise ValueError("no_orders")

    wb = Workbook()
    ws_orders = wb.active
    ws_orders.title = "Заказы"
    ws_orders.append(["Пользователь", "Блюдо", "Столовая", "Время"])

    summary: dict[str, int] = {}
    for order in orders:
        user_name = order.user.full_name if order.user else f"tg:{order.user_id}"
        meal_title = order.meal.title if order.meal else f"meal:{order.meal_id}"
        canteen_title = order.meal.canteen.title if order.meal and order.meal.canteen else "—"
        tz = ZoneInfo(settings.timezone)
        local_time = order.created_at.replace(tzinfo=timezone.utc).astimezone(tz)
        ws_orders.append([
            user_name,
            meal_title,
            canteen_title,
            local_time.strftime("%H:%M"),
        ])
        summary[meal_title] = summary.get(meal_title, 0) + 1

    ws_summary = wb.create_sheet("Сводка")
    ws_summary.append(["Блюдо", "Кол-во"])
    for meal_title, count in summary.items():
        ws_summary.append([meal_title, count])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ts = date.strftime("%d.%m_%H-%M")
    return f"Отчёт_{ts}.xlsx", buffer


# ======== Отчёт по конкретному меню ========


async def build_menu_report(menu_id: int) -> Tuple[str, BytesIO]:
    """Создаёт Excel-отчёт только по заказам данного меню (DailyMenu.id)."""
    async with get_session() as session:
        # Определяем список menu_id: основной и все копии с parent_id = menu_id
        menus_res = await session.exec(select(MenuItem.menu_id))
        child_res = await session.exec(select(DailyMenu.id).where(DailyMenu.parent_id == menu_id))
        menu_ids = [menu_id] + list(child_res)

        stmt = (
            select(Order)
            .where(Order.menu_id.in_(menu_ids), Order.is_final == True)  # noqa: E712
            .options(
                selectinload(Order.user),
                selectinload(Order.meal).selectinload(Meal.canteen),
            )
        )
        orders = list(await session.exec(stmt))

    if not orders:
        raise ValueError("no_orders")

    wb = Workbook()
    ws_orders = wb.active
    ws_orders.title = "Заказы"
    ws_orders.append(["Пользователь", "Блюдо", "Столовая", "Время"])

    summary: dict[str, int] = {}
    for order in orders:
        user_name = order.user.full_name if order.user else f"tg:{order.user_id}"
        meal_title = order.meal.title if order.meal else f"meal:{order.meal_id}"
        canteen_title = order.meal.canteen.title if order.meal and order.meal.canteen else "—"
        tz = ZoneInfo(settings.timezone)
        local_time = order.created_at.replace(tzinfo=timezone.utc).astimezone(tz)
        ws_orders.append([
            user_name,
            meal_title,
            canteen_title,
            local_time.strftime("%H:%M"),
        ])
        summary[meal_title] = summary.get(meal_title, 0) + 1

    ws_summary = wb.create_sheet("Сводка")
    ws_summary.append(["Блюдо", "Кол-во"])
    for meal_title, count in summary.items():
        ws_summary.append([meal_title, count])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ts = datetime.now().strftime("%d.%m_%H-%M")
    return f"Отчёт_меню_{ts}.xlsx", buffer 